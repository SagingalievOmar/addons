bl_info = {
    "name": "Furniture Panel Calculator",
    "author": "Omar",
    "version": (2, 3),
    "blender": (3, 0, 0),
    "location": "View3D > Tool Shelf > Panel Calculator",
    "description": "Automatically save panel dimensions and edge banding for each selected object individually",
    "category": "Object",
}

import bpy
import os
from openpyxl import Workbook, load_workbook


class EdgeBandingProperties(bpy.types.PropertyGroup):
    edge_top: bpy.props.BoolProperty(name="Top (Longest Side)", default=False) # type: ignore
    edge_right: bpy.props.BoolProperty(name="Right Side", default=False) # type: ignore
    edge_bottom: bpy.props.BoolProperty(name="Bottom Side", default=False) # type: ignore
    edge_left: bpy.props.BoolProperty(name="Left Side", default=False) # type: ignore


class PANEL_PT_PanelCalculator(bpy.types.Panel):
    """Creates a Panel in the Object properties window"""
    bl_label = "Panel Calculator"
    bl_idname = "PANEL_PT_panel_calculator"
    bl_space_type = 'VIEW_3D'
    bl_region_type = 'UI'
    bl_category = 'Panel Calculator'

    def draw(self, context):
        layout = self.layout

        layout.operator("object.open_edge_banding_window", text="Select Edge Banding")

        layout.operator("object.save_edge_banding_data", text="Save Edge Banding Data")


class OBJECT_OT_OpenEdgeBandingWindow(bpy.types.Operator):
    """Open a window to mark edge banding sides"""
    bl_idname = "object.open_edge_banding_window"
    bl_label = "Edge Banding Options"

    def invoke(self, context, event):
        return context.window_manager.invoke_props_dialog(self)

    def draw(self, context):
        layout = self.layout
        layout.label(text="Mark Edge Banding for Selected Panels")

        row = layout.row()
        row.label(text="Mass Selection:")
        row = layout.row()
        row.operator("object.set_mass_edge_banding", text="Top").side = "TOP"
        row.operator("object.set_mass_edge_banding", text="Right").side = "RIGHT"
        row.operator("object.set_mass_edge_banding", text="Bottom").side = "BOTTOM"
        row.operator("object.set_mass_edge_banding", text="Left").side = "LEFT"
        
        row = layout.row()
        row.label(text="Mass Deselection:")
        row = layout.row()
        row.operator("object.unset_mass_edge_banding", text="Unset Top").side = "TOP"
        row.operator("object.unset_mass_edge_banding", text="Unset Right").side = "RIGHT"
        row.operator("object.unset_mass_edge_banding", text="Unset Bottom").side = "BOTTOM"
        row.operator("object.unset_mass_edge_banding", text="Unset Left").side = "LEFT"

        for obj in context.selected_objects:
            dim = list(map(lambda x: round(x * 1000), sorted(obj.dimensions)))
            if obj.type == 'MESH' and 9 < dim[0] < 33:
                layout.label(text=f"{obj.name}: {dim[2]}x{dim[1]}")
                props = obj.edge_banding_props

                layout.prop(props, "edge_top")
                layout.prop(props, "edge_right")
                layout.prop(props, "edge_bottom")
                layout.prop(props, "edge_left")
            else:
                self.report({"INFO"}, f"Объект {obj.name} не является деталью для кромковки!")
                continue

    def execute(self, context):
        self.report({'INFO'}, "Edge banding updated for selected objects.")
        return {'FINISHED'}


class OBJECT_OT_SetMassEdgeBanding(bpy.types.Operator):
    """Set edge banding for all objects"""
    bl_idname = "object.set_mass_edge_banding"
    bl_label = "Set Mass Edge Banding"

    side: bpy.props.StringProperty() # type: ignore

    def execute(self, context):
        for obj in context.selected_objects:
            dim = list(map(lambda x: round(x * 1000), sorted(obj.dimensions)))
            if obj.type == 'MESH' and hasattr(obj, 'edge_banding_props') and 9 < dim[0] < 33:
                props = obj.edge_banding_props
                if self.side == "TOP":
                    props.edge_top = True
                elif self.side == "RIGHT":
                    props.edge_right = True
                elif self.side == "BOTTOM":
                    props.edge_bottom = True
                elif self.side == "LEFT":
                    props.edge_left = True

        self.report({'INFO'}, f"Set {self.side} edge banding for all objects.")
        return {'FINISHED'}

class OBJECT_OT_UnsetMassEdgeBanding(bpy.types.Operator):
    """Unset edge banding for all objects"""
    bl_idname = "object.unset_mass_edge_banding"
    bl_label = "Unset Mass Edge Banding"

    side: bpy.props.StringProperty() # type: ignore

    def execute(self, context):
        for obj in context.selected_objects:
            dim = list(map(lambda x: round(x * 1000), sorted(obj.dimensions)))
            if obj.type == 'MESH' and hasattr(obj, 'edge_banding_props') and 9 < dim[0] < 33:
                props = obj.edge_banding_props
                if self.side == "TOP":
                    props.edge_top = False
                elif self.side == "RIGHT":
                    props.edge_right = False
                elif self.side == "BOTTOM":
                    props.edge_bottom = False
                elif self.side == "LEFT":
                    props.edge_left = False
            

        self.report({'INFO'}, f"Unset {self.side} edge banding for all objects.")
        return {'FINISHED'}

class OBJECT_OT_SaveEdgeBandingData(bpy.types.Operator):
    """Save edge banding data without overwriting existing data"""
    bl_idname = "object.save_edge_banding_data"
    bl_label = "Save Edge Banding Data"

    def execute(self, context):
        blend_file_path = bpy.data.filepath
        if not blend_file_path:
            self.report({'ERROR'}, "Please save the Blender file before exporting data.")
            return {'CANCELLED'}

        base_name = os.path.splitext(os.path.basename(blend_file_path))[0]
        file_path = os.path.join(os.path.dirname(blend_file_path), f"{base_name}_ldsp.xlsx")
        file_path_hdf = os.path.join(os.path.dirname(blend_file_path), f"{base_name}_hdf.xlsx")

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Name", "Type", "Length", "Width", "Quantity",
                       "Top", "Right", "Bottom", "Left"]
            sheet.append(headers)
            
        if os.path.exists(file_path_hdf):
            workbook_hdf = load_workbook(file_path_hdf)
            sheet_hdf = workbook_hdf.active
        else:
            workbook_hdf = Workbook()
            sheet_hdf = workbook_hdf.active
            headers_hdf = ["Name", "Type", "Length", "Width", "Quantity"]
            sheet_hdf.append(headers_hdf)
            
        information = []
        for obj in context.selected_objects:
            inf = (obj.location, obj.dimensions)
            if inf not in information:
                information.append(inf)
            else:
                self.report({'ERROR'}, f"Объект '{obj.name}' перекрывает объект другой обьект")
                return {'CANCELLED'}

        for obj in context.selected_objects:
            if obj.type == 'MESH':
                dimensions = sorted(obj.dimensions, reverse=True)
                length, width, thickness = [round(dim * 1000) for dim in dimensions]

                material_type = "LDSP" if 9 < thickness < 21 else "HDF" if 2 < thickness < 5 else "Unknown"
                if obj.name[:4] != "Cube":
                    self.report({'ERROR'}, f"объект {obj.name} был пропущен так как он не является обрабатываемой деталью БУДЬ ВНИМАТЕЛЕН!!!")
                    continue
                
                if material_type is "HDF":
                    new_data_hdf = [
                        obj.name,
                        material_type,
                        length,
                        width,
                        1,
                        ("", 0.5)[obj.edge_banding_props.edge_top]
                    ]
                    
                    for row in sheet_hdf.iter_rows(min_row=2, max_row=sheet_hdf.max_row, values_only=False):
                        if row[0].value == obj.name:
                            for col, value in enumerate(new_data_hdf, start=1):
                                row[col - 1].value = value
                            break
                    else:
                        sheet_hdf.append(new_data_hdf)
                else:
                    new_data = [
                        obj.name,
                        material_type,
                        length,
                        width,
                        1,
                        ("", 0.5)[obj.edge_banding_props.edge_top],
                        ("", 0.5)[obj.edge_banding_props.edge_right],
                        ("", 0.5)[obj.edge_banding_props.edge_bottom],
                        ("", 0.5)[obj.edge_banding_props.edge_left],
                    ]
                    
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                        if row[0].value == obj.name:
                            for col, value in enumerate(new_data, start=1):
                                row[col - 1].value = value
                            break
                    else:
                        sheet.append(new_data)
                
        workbook_hdf.save(file_path_hdf)
        workbook.save(file_path)
        self.report({'INFO'}, f"Saved edge banding data to {file_path}.")
        self.report({"INFO"}, f"Saved hdf data to {file_path_hdf}.")
        return {'FINISHED'}


def register():
    bpy.utils.register_class(EdgeBandingProperties)
    bpy.types.Object.edge_banding_props = bpy.props.PointerProperty(type=EdgeBandingProperties)

    bpy.utils.register_class(PANEL_PT_PanelCalculator)
    bpy.utils.register_class(OBJECT_OT_OpenEdgeBandingWindow)
    bpy.utils.register_class(OBJECT_OT_SetMassEdgeBanding)
    bpy.utils.register_class(OBJECT_OT_UnsetMassEdgeBanding)
    bpy.utils.register_class(OBJECT_OT_SaveEdgeBandingData)


def unregister():
    bpy.utils.unregister_class(EdgeBandingProperties)
    del bpy.types.Object.edge_banding_props

    bpy.utils.unregister_class(PANEL_PT_PanelCalculator)
    bpy.utils.unregister_class(OBJECT_OT_OpenEdgeBandingWindow)
    bpy.utils.unregister_class(OBJECT_OT_SetMassEdgeBanding)
    bpy.utils.register_class(OBJECT_OT_UnsetMassEdgeBanding)
    bpy.utils.unregister_class(OBJECT_OT_SaveEdgeBandingData)


if __name__ == "__main__":
    register()
